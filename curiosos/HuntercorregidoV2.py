from pyhunter import PyHunter
from openpyxl import Workbook

# Alumno: Enrique Adrian Gonzalez Martinez
# Alumno: Enrique de Jesus Vazquez Perez
# Los comentarios deben de tener un espacio al inicio
# Sustituye la siguiente API key por la tuya
# Al darle un valor a una funcion no hay que dejar ningun espacio
hunter = PyHunter('47bb2e9489599c41941328d6a0c5ddcd53a31115')


# Se encarga de buscar la organizacion que se ingrese
def busqueda(organizacion):
    # Hay que dejar 4 espacios de indentado
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 50, cuidado!
    busqueda = 1
    # Las lineas de codigo no pueden exceder los 79 caracteres
    resultado = hunter.domain_search(company=organizacion,
                                     limit=busqueda,
                                     emails_type='personal')
    return resultado


# Se encarga de guardar los datos de la organizacion y guardarlos en un excel
def guardar_info(datos_encontrados, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    hoja.cell(1, 1, "Dominio")
    hoja.cell(1, 2, datos_encontrados['domain'])
    hoja.cell(2, 1, "Organización")
    hoja.cell(2, 2, datos_encontrados['organization'])
    hoja.cell(3, 1, "Correo")
    hoja.cell(3, 2, datos_encontrados['emails'][0]['value'])
    hoja.cell(4, 1, "Nombre(s)")
    hoja.cell(4, 2, datos_encontrados['emails'][0]['first_name'])
    hoja.cell(5, 1, "Apellidos")
    hoja.cell(5, 2, datos_encontrados['emails'][0]['last_name'])
    libro.save("Hunter" + organizacion + ".xlsx")


print("Script para buscar información")
orga = input("Organización a investigar: ")
datos_encontrados = busqueda(orga)
# Si la comparacion es None deberia ser "if cond is None"
if datos_encontrados is None:
    exit()
else:
    print(datos_encontrados)
    print(type(datos_encontrados))
    guardar_info(datos_encontrados, orga)
